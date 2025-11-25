"""
Export test cases to Excel spreadsheet (.xlsx)
Requires: openpyxl
Install: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_test_case_excel(test_cases, output_file="test_cases.xlsx"):
    """
    Convert test cases to a formatted Excel spreadsheet
    
    Args:
        test_cases: List of dictionaries containing test case information
        output_file: Output filename for the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define headers
    headers = [
        'Test Case ID',
        'Test Case Name',
        'Description',
        'Preconditions',
        'Test Steps',
        'Expected Result',
        'Actual Result',
        'Status',
        'Priority',
        'Test Type'
    ]
    
    # Add title row
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = 'Test Cases Documentation'
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Add metadata row
    ws.merge_cells('A2:J2')
    metadata_cell = ws['A2']
    metadata_cell.value = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total Test Cases: {len(test_cases)}'
    metadata_cell.font = Font(size=10, italic=True)
    metadata_cell.alignment = Alignment(horizontal='center')
    
    # Add headers
    header_row = 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    ws.row_dimensions[header_row].height = 35
    
    # Add test case data
    for row_num, test_case in enumerate(test_cases, start=header_row + 1):
        data = [
            test_case.get('id', f'TC_{row_num - header_row:03d}'),
            test_case.get('name', 'Unnamed Test'),
            test_case.get('description', 'N/A'),
            test_case.get('preconditions', 'N/A'),
            test_case.get('steps', 'N/A'),
            test_case.get('expected_result', 'N/A'),
            test_case.get('actual_result', ''),
            test_case.get('status', 'Not Executed'),
            test_case.get('priority', 'Medium'),
            test_case.get('test_type', 'Functional')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Status color coding
            if col_num == 8:  # Status column
                if value.lower() == 'pass':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                elif value.lower() == 'fail':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
        
        ws.row_dimensions[row_num].height = 60
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # ID
        'B': 30,  # Name
        'C': 35,  # Description
        'D': 25,  # Preconditions
        'E': 40,  # Steps
        'F': 35,  # Expected
        'G': 35,  # Actual
        'H': 15,  # Status
        'I': 12,  # Priority
        'J': 15   # Type
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save workbook
    wb.save(output_file)
    print(f"✓ Excel file saved: {output_file}")
    return output_file


# Example usage
if __name__ == "__main__":
    # Sample test cases structure
    sample_test_cases = [
        {
            "id": "TC_001",
            "name": "User Login with Valid Credentials",
            "description": "Verify that user can login with valid username and password",
            "preconditions": "User account exists in the system",
            "steps": "1. Navigate to login page\n2. Enter valid username\n3. Enter valid password\n4. Click Login button",
            "expected_result": "User is successfully logged in and redirected to dashboard",
            "status": "Pass",
            "priority": "High",
            "test_type": "Functional"
        },
        {
            "id": "TC_002",
            "name": "User Login with Invalid Password",
            "description": "Verify error message when invalid password is entered",
            "preconditions": "User account exists in the system",
            "steps": "1. Navigate to login page\n2. Enter valid username\n3. Enter invalid password\n4. Click Login button",
            "expected_result": "Error message 'Invalid credentials' is displayed",
            "status": "Fail",
            "priority": "High",
            "test_type": "Negative Testing"
        }
    ]
    
    # Create Excel file
    create_test_case_excel(sample_test_cases, "test_cases_output.xlsx")
